import json
import logging
from typing import List

from langchain.document_loaders.base import BaseLoader
from langchain.schema import Document
from openpyxl.reader.excel import load_workbook

logger = logging.getLogger(__name__)


class ExcelLoader(BaseLoader):
    """Load xlxs files.


    Args:
        file_path: Path to the file to load.
    """

    def __init__(
        self,
        file_path: str
    ):
        """Initialize with file path."""
        self._file_path = file_path

    def load(self) -> List[Document]:
        data = []
        keys = []
        wb = load_workbook(filename=self._file_path, read_only=True)
        # loop over all sheets
        for sheet in wb:
            for row in sheet.iter_rows(values_only=True):
                if all(v is None for v in row):
                    continue
                if keys == []:
                    keys = list(map(str, row))
                else:
                    row_dict = dict(zip(keys, row))
                    row_dict = {k: v for k, v in row_dict.items() if v}
                    data.append(json.dumps(row_dict, ensure_ascii=False))

        return [Document(page_content='\n\n'.join(data))]
